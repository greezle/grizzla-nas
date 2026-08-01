"""Filtered failure exports: CSV for quick looks, XLSX for the farm's
Excel workflows. Both take the same URL query as the browser page, so the
export always matches what the filters show."""
import csv
import io
import time

from awaria.services.failures import failures_select

HEADER = [
    "ID", "Drukarka", "Kategoria", "Blokada", "Otwarta", "Naprawiona",
    "Czas [h]", "Zamknięta przez", "Podczas wydruku", "Szczegóły",
    "Notatka serwisowa", "Komentarze"
]


def _rows(db, query):
    now = int(time.time())
    for f in failures_select(db, query):
        opened = f["opened_ts"] or now
        hours = round(((f["closed_ts"] or now) - opened) / 3600, 1)
        yield [
            f["id"], f["hostname"], f["label"] or "",
            "TAK" if f["blocking"] else "NIE", f["opened_at"] or "",
            f["closed_at"] or "", hours, f["closed_by"] or "",
            f["print_file"] or "", f["detail"] or "", f["repair_note"] or "",
            f["comments_joined"] or ""
        ]


def export_failures_csv(db, query):
    """Semicolon separator + BOM + comma decimals: what Polish Excel expects
    when double-clicking a .csv."""
    out = io.StringIO()
    writer = csv.writer(out, delimiter=";", lineterminator="\r\n")
    writer.writerow(HEADER)
    for row in _rows(db, query):
        row[6] = str(row[6]).replace(".", ",")
        writer.writerow(row)
    return "﻿" + out.getvalue()


PRINTS_HEADER = [
    "Drukarka", "Plik", "Start", "Koniec", "Czas [h]", "Szacowany czas [h]",
    "Wynik", "Rodzaj", "Filament", "Podkładka"
]

RESULT_TEXT = {
    "finished": "ukończony",
    "finished?": "ukończony",
    "aborted": "anulowany",
    "aborted?": "anulowany",
}


# Rows stream from a cursor into a write-only workbook, so memory barely
# tracks the row count: measured on the Pi, 100k rows cost +3.5 MB of RSS
# and a 3 MB file (a real export runs ~60 B/row in the file). Memory is
# therefore not the binding constraint - this cap exists to bound how long
# the export holds the database lock, and 200k rows is roughly two years of
# the whole farm.
EXPORT_MAX_ROWS = 200000


def export_prints_xlsx(db, query):
    """The Historia table as a workbook - same filters, same rows, without
    the on-screen display cap. Returns None when openpyxl is missing, or the
    row count (an int) when the selection is too large to build safely."""
    try:
        import openpyxl
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None
    # imported here: pages imports nothing from this module, so this stays a
    # one-way dependency
    from awaria.web.pages import (gcode_meta_index, meta_of_print,
                                  prints_count, prints_iter)

    total = prints_count(db, query)
    if total > EXPORT_MAX_ROWS:
        return total

    meta_idx = gcode_meta_index(db)
    now = int(time.time())
    # write-only mode streams rows out instead of holding a cell object per
    # value - the difference that keeps a big export inside the memory cap
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Wydruki")
    for i, width in enumerate([10, 52, 19, 19, 9, 17, 11, 9, 14, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    header = []
    for title in PRINTS_HEADER:
        cell = WriteOnlyCell(ws, value=title)
        cell.font = Font(bold=True)
        header.append(cell)
    ws.append(header)
    for p in prints_iter(db, query):
        meta = meta_of_print(meta_idx, p["file"])
        hours = round(((p["ended_ts"] or now) - (p["started_ts"] or now)) /
                      3600, 2)
        ws.append([
            p["hostname"], p["file"], p["started_at"], p["ended_at"] or "",
            hours,
            round(meta["est_s"] / 3600, 2) if meta and meta["est_s"] else "",
            "w trakcie" if not p["ended_at"] else RESULT_TEXT.get(
                p["result"], "nieznany"), p["kind"],
            (meta["filament"] if meta else None) or p["material"] or "",
            (meta["sheet"] if meta else "") or ""
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_failures_xlsx(db, query):
    """Returns the workbook bytes, or None when openpyxl is unavailable
    (it is on the NAS - the g-code publisher already depends on it)."""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Awarie"
    ws.append(HEADER)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for row in _rows(db, query):
        ws.append(row)
    for i, width in enumerate([6, 10, 24, 9, 17, 17, 8, 13, 28, 40, 30, 40],
                              start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
